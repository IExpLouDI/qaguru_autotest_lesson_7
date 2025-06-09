from openpyxl import load_workbook
from utils.paths import DIR_WITH_RESOURCES
import os


if __name__ == "main":
    workbook = load_workbook(os.path.join(DIR_WITH_RESOURCES, "file_example_XLSX_50.xlsx"))
    sheet = workbook.active # получаем активный лист

    print(sheet.cell(row=3, column=2).value) # печатаем содержимое строки и колонки

# for row in sheet.iter_rows(): # получаем все строки и печатаем их
#     print(row)
#
# for column in sheet.iter_cols(): # получаем все колонки и печатаем их
#     print(column)
#
# for row in sheet.iter_rows():  # печатаем содержимое всех ячеек
#     for cell in row:
#         print(cell.value)
