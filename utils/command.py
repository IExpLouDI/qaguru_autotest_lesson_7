import os
from xlrd import open_workbook
from openpyxl import load_workbook
from pypdf import PdfReader
from source.script_xls import FileInfoXls


def check_result(f_type: str, instruction: dict, _object):
	keys = instruction.keys()

	if f_type == ".xls":
		book = open_workbook(file_contents=_object)
		info_xls = FileInfoXls(book)
		res_dict = info_xls.get_file_info()
	elif f_type == ".xlsx":
		book = load_workbook(_object)
		sheet = book.active
		res_dict = {"value_3_2": sheet.cell(row=3, column=2).value}
	elif f_type == ".pdf":
		reader = PdfReader(_object)
		res_dict = {
			"page_count": len(reader.pages)
		}

	for key in keys:
		if instruction[key] != res_dict[key]:
			return [False, f"\nBad value - {key} in file type - {f_type}."
						   f"\nExpected - {instruction[key]}, but - {res_dict[key]} given!"]

	return [True, "Congratulations"]
